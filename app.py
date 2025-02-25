import pandas as pd
import flet as ft
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
import os
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import locale
import re
import xmltodict

locale.setlocale(locale.LC_TIME, "Portuguese_Brazil.1252")

datepicker_aberto = False
area_de_trabalho = os.path.expanduser("~/Desktop")
caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")
caminhoDaPasta = ""

def main(page : ft.Page):
    area_de_trabalho = os.path.expanduser("~/Desktop")
    caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")
    placas = ['IJN2169', 'IPA1605', 'IQH4554', 'IQQ8275', 'IQU5856', 'IRT4484', 'IRT6007', 'IRT8903', 'IRT8908', 'ISD6797', 'ISD6799', 'ISW4018', 'ISW4036', 'ITR9934', 'ITR9987', 'IUK1811', 'IUL0815', 'IUN1285', 'IUQ6I14', 'IUU2781', 'IUU2790', 'IUU2976', 'IUY5953', 'IVD1754', 'IVD1767 ', 'IVD2882', 'IVD2886', 'IVD2894', 'IUU2766  ', 'ISW4049  ', 'IPH0196  ', 'ISC7987', 'IQQ8279', 'IUK1654', 'IQQ8280', 'IQX1137', 'IRS6482', 'IRU1028', 'ISC7522', 'IXU8288', 'IVS9811', 'IVV7697', 'IVZ5017', 'IXM3552 ', 'IYM9529 ', 'IXN5003', 'IYV6I97', 'IYW1850', 'IZY9A38', 'IZZ3B77', 'IZZ7E85', 'JAA7G36', 'JAA9A85', 'JAV8C49', 'JAW2H90', 'JBH1F60', 'JBT7G53', 'JBT9A56', 'JBU9F23', 'JBV5I08', 'JBW4A65', 'JBX9J56', 'JCA6H69', 'JCJ9F30', 'JCM3A76', 'JCN1F03', 'JCW0B99', 'IUP9926', 'IWF4461', 'IWP3711', 'IWR7894', 'IWT8079', 'IXP2159', 'IXS9707', 'IYI1674', 'IYT9710', 'IYU3191', 'IYY4885', 'IYZ7G33', 'IZC5G49', 'IZC9F27', 'IYU7F23', 'IZF9F66', 'ISS3661', 'IUD5253', 'IWI2587', 'ISD6653', 'ISU2509', 'IUE3232', 'IUC5395', 'IUQ0086', 'ITZ7769', 'IUA1579', 'IQS4075', 'ITA3044', 'ISX9718 ', 'IXO2988 ', 'IYZ9H14 ', 'IZG4J25 ', 'IZS2E90 ', 'JAJ1J25 ', 'JBH7E56', 'JBJ6A76 ', 'JCD6F25', 'JCI2B18', 'IPU2889', 'ISW6246', 'IUD6684', 'IUO4697', 'IWE1022', 'JCC6C61', 'IWD6417', 'IXM3924', 'IXM5236', 'IXU4276', 'IYE9915', 'IZE9G51', 'IZI0J48', 'IZJ3J83', 'IZJ9E19', 'IZK5A16', 'IZK5A30', 'IYF3334', 'IZY9A82', 'JCF6C26', 'JCF6C25', 'IXA3764', 'ITI0766', 'ISZ9602', 'IVR8722', 'ISQ3832', 'ITB2764', 'ITB2776', 'IUO2076 ', 'ONO8412', 'ONO9432', 'IWT6897', 'ONO8622', 'IUL6837', 'IVM6958', 'IVQ4C15', 'IYA3499', 'IYE3493', 'IYE3522', 'IYJ0117', 'IYJ3922', 'IYZ8E65', 'IZB3E36', 'IZO3J10', 'IZR5H25', 'JAA0C41', 'JBI6C59', 'JBK6C92', 'JBR8G27', 'JBT6G13', 'JBU3B31', 'JBV9H85', 'JBV9H40', 'JCG4C96', 'JCV0A51', 'JCU9J11', 'IHV1577', 'IKJ3953', 'IPI0196', 'IQJ0890', 'IQL9597', 'IQU8661', 'IQU8666', 'IQU8668', 'IQU3167  ', 'IQH2603 ', 'IWX8573 ', 'IZK8E80', 'IZK8E86', 'IZX3B56 ', 'IZX3I40', 'IZX3C03', 'IZX3I48', 'IZX3B28', 'JAE9F27', 'JAF0D41', 'JAF7B20', 'JAH2G15', 'JAF4G65', 'JAF6E08', 'JAH6C84', 'JAH5A67', 'JAH7D78', 'JAH4G61', 'JBG0G17', 'JBG0E53', 'JBG0F36', 'JBG7J23', 'JBH5G30', 'JBS0B74', 'JBS0B75', 'JBS0B76', 'JBS9D03', 'JBS9C26', 'JCS6C77', 'JCS6C37', 'JCU4G00', 'JCS9J25', 'JCS9J44', 'JCV3B16', 'JCV3A96', 'JCV6G29', 'JCW6G83', 'JCW6G98', 'IYA4653', 'IEN5388']
    pb = ft.ProgressBar(width=200)
    pbText = ft.Text("Iniciando processo")
    ColunaPb = ft.Column([pb, pbText], alignment=ft.alignment.center, horizontal_alignment=ft.CrossAxisAlignment.CENTER, visible=False)
    
    
    
    
    
    if not os.path.exists(caminho_arquivo):
            # Criar uma nova planilha
            wb = Workbook()
            wb.remove(wb.active)  # Remover a aba padrão "Sheet"
            
            # Estilo para cabeçalho
            negrito = Font(bold=True)
            borda = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Criar abas para cada placa e adicionar os cabeçalhos
            for placa in placas:
                ws = wb.create_sheet(title=placa)
                headers = ["PLACA", "DATA", "POSTO", "KM", "LITRAGEM", "MEDIA"]
                ws.append(headers)
                
                # Aplicar estilos ao cabeçalho
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = negrito
                    cell.border = borda
            
            # Salvar a planilha
            wb.save(caminho_arquivo)
        
        
    page.title = "Combustiveis"
    page.window.width = 700
    page.window.height = 500
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    listaPostos = ["S I M", "BOHRER", "TRADIÇÃO (NÃO DISPONIVEL)", "BUFFON", "SANDER (POSTO SUSPENSO)"]
    postos = ft.Dropdown(
        label="Postos",
        options=[ft.dropdown.Option(posto) for posto in listaPostos],
        value=listaPostos[0],
        on_change=lambda e: page.update(),
    )

        
    def sim(dataInicial, dataFinal):
        
        ColunaPb.visible = True

        pb.value = 0
        page.update()
        
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.get("https://portalb2b.simrede.com.br/login")
        driver.maximize_window()
        wait = WebDriverWait(driver, 500)
        action = ActionChains(driver)
        time.sleep(1)
        pbText.value = "Realizando login"
        pb.value = 0.1
        page.update()
        userLogin = 'manutencao.rede@gmail.com'
        senhaLogin = 'noll3020'
        campoLogin = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/app-auth/div/div/div/div/div/div/form/div[1]/input')))
        campoSenha = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/app-auth/div/div/div/div/div/div/form/div[2]/input')))
        campoLogin.send_keys(userLogin)
        campoSenha.send_keys(senhaLogin)
        campoSenha.send_keys(Keys.RETURN)
        time.sleep(17)
        pbText.value = "Realizando consulta"
        pb.value = 0.3
        page.update()
        dropDownRelatorios = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="side-nav-horizontal"]/app-nav-group/app-nav-collapse[3]/li/a' )))
        dropDownRelatorios.click()
        menuCompraDiaria = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="side-nav-horizontal"]/app-nav-group/app-nav-collapse[3]/li/ul/app-nav-item[2]/li/a')))
        menuCompraDiaria.click()
        time.sleep(1)
        dataInicialFiltro = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltros"]/div/app-loading/div/div/form/div[4]/div[3]/div/input')))
        dataInicialFiltro.send_keys(dataInicial)
        dataFinalFiltro = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltros"]/div/app-loading/div/div/form/div[4]/div[4]/div/input')))
        dataFinalFiltro.send_keys(dataFinal)
        botaoConsultar = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltros"]/div/app-loading/div/div/form/div[5]/div/button[1]')))
        botaoConsultar.click()
        pbText.value = "Realizando download da planilha"
        pb.value = 0.5
        page.update()
        botaoRelExcel = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonDownload"]')))
        botaoRelExcel.click()


        def abrir_ultima_planilha_downloads():
            pasta_downloads = os.path.expanduser("~/Downloads")
            arquivos_planilhas = glob.glob(os.path.join(pasta_downloads, "*.xls*"))
            
            if not arquivos_planilhas:
                return None, None
            
            ultima_planilha = max(arquivos_planilhas, key=os.path.getctime)
            
            try:
                df = pd.read_excel(ultima_planilha)
                return df, ultima_planilha
            except Exception as e:
                return None, None

        def criar_ou_atualizar_planilha(df):
            try:
                df = df[~df.iloc[:, 15].astype(str).str.contains("ARLA", na=False, regex=True)]
                
                colunas_desejadas = [4, 16, 17, 11]  
                df_selecionado = df.iloc[:, colunas_desejadas].copy()
                df_selecionado.columns = ["PLACA", "DATA", "KM", "LITRAGEM"]
                df_selecionado.insert(2, "POSTO", "S I M")
                df_selecionado.sort_values(by=["PLACA", "DATA"], inplace=True)
                
                df_selecionado["KM"] = pd.to_numeric(df_selecionado["KM"], errors="coerce")
                df_selecionado["LITRAGEM"] = pd.to_numeric(df_selecionado["LITRAGEM"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
                df_selecionado["DATA"] = pd.to_datetime(df_selecionado["DATA"], errors="coerce").dt.strftime('%d/%m/%Y')

                area_de_trabalho = os.path.expanduser("~/Desktop")
                caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")

                if os.path.exists(caminho_arquivo):
                    book = load_workbook(caminho_arquivo)
                else:
                    book = None

                placas_unicas = df_selecionado["PLACA"].unique()

                with pd.ExcelWriter(caminho_arquivo, engine="openpyxl", mode="a" if book else "w", if_sheet_exists="replace") as writer:
                    for placa in placas_unicas:
                        df_placa = df_selecionado[df_selecionado["PLACA"] == placa].copy()
                        df_placa = df_placa.sort_values(by="DATA")

                        if book and placa in book.sheetnames:
                            existing_df = pd.read_excel(caminho_arquivo, sheet_name=str(placa), dtype=str)
                            existing_df["DATA"] = pd.to_datetime(existing_df["DATA"], format="%d/%m/%Y", errors="coerce").dt.strftime('%d/%m/%Y')

                            df_placa["DATA"] = pd.to_datetime(df_placa["DATA"], format="%d/%m/%Y", errors="coerce").dt.strftime('%d/%m/%Y')

                            # Unificar formatos de colunas antes da comparação
                            existing_df["KM"] = pd.to_numeric(existing_df["KM"], errors="coerce")
                            existing_df["LITRAGEM"] = pd.to_numeric(existing_df["LITRAGEM"], errors="coerce")
                            
                            df_placa["KM"] = pd.to_numeric(df_placa["KM"], errors="coerce")
                            df_placa["LITRAGEM"] = pd.to_numeric(df_placa["LITRAGEM"], errors="coerce")

                            # Encontrar registros novos
                            df_placa = df_placa.merge(
                                existing_df, on=["PLACA", "DATA", "KM", "POSTO", "LITRAGEM"], how="left", indicator=True
                            )
                            df_placa = df_placa[df_placa["_merge"] == "left_only"].drop(columns=["_merge"])

                            df_placa = pd.concat([existing_df, df_placa], ignore_index=True)
                            df_placa.drop_duplicates(subset=["PLACA", "DATA", "KM", "POSTO", "LITRAGEM"], keep="first", inplace=True)
                            df_placa = df_placa.sort_values(by="DATA")

                        df_placa.to_excel(writer, sheet_name=str(placa), index=False)

                book = load_workbook(caminho_arquivo)
                
                for placa in placas_unicas:
                    sheet = book[str(placa)]
                    max_row = sheet.max_row

                    for row in range(2, max_row + 1):
                        if row > 2:
                            formula = f'=IF(A{row}<>"",(D{row}-D{row-1})/E{row},"")'
                            sheet[f"F{row}"] = formula

                book.save(caminho_arquivo)

            except Exception as e:
                print(f"Erro ao criar ou atualizar a planilha: {e}")
        
        
        time.sleep(3)
        df, ultima_planilha = abrir_ultima_planilha_downloads()
        if df is not None:
            pb.value = 0.8
            pbText.value = "Criando/atualizando planilha"
            page.update()
            
            criar_ou_atualizar_planilha(df)
            pb.value = 1
            pbText.value = "Planilha criada/atualizada com sucesso"
            page.update()
            time.sleep(1)
            pb.value = 0
            pbText.value = "Iniciando processo"
            ColunaPb.visible = False
            page.update()
        time.sleep(1)
        
        driver.quit()
        return

        
    def bohrer(dataInicial, dataFinal):
        ColunaPb.visible = True

        pb.value = 0
        page.update()
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.get("https://www.qualityautomacao.com.br/webfrota/#/login")
        driver.maximize_window()
        wait = WebDriverWait(driver, 500)
        pbText.value = "Realizando login"
        pb.value = 0.1
        page.update()
        
        userLogin = 'Redeencomendas'
        senhaLogin = '1945Luciano'
        campoLogin = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/app-root/app-login/main/div/div[2]/cla-card/div/cla-card-body/div/form/div/div[1]/cla-input/fieldset/input')))
        campoSenha = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/app-root/app-login/main/div/div[2]/cla-card/div/cla-card-body/div/form/div/div[2]/cla-input-password/fieldset/input')))
        campoLogin.send_keys(userLogin)
        campoSenha.send_keys(senhaLogin)
        campoSenha.send_keys(Keys.RETURN)
        time.sleep(4)
        pbText.value = "Realizando consulta"
        pb.value = 0.3
        page.update()
        relatorio = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menu"]/nav/app-sidebar/ul/li[7]/a')))
        relatorio.click()
        historicoConsumo = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="relatorio"]/ul/li[1]/a')))
        historicoConsumo.click()
        filtroData = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mobile"]/main/app-consumption-history/app-consumption-history-card-title/section/div[1]/fieldset/button')))
        filtroData.click()
        campoData = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="history_filter"]/div[2]/form/div/div[5]/cla-input-date/fieldset/button')))
        campoData.click()
        time.sleep(2)
        
        
        
        #INICIAL
        def remover_celulas_anteriores_a_um(celulas):
            for i, celula in enumerate(celulas):
                if celula.text.isdigit() and int(celula.text) == 1:
                    return celulas[i:]
            return []
        mesHoje = datetime.now().strftime("%b. de %Y").lower()
        dataInicial_obj = datetime.strptime(str(dataInicial), "%d/%m/%Y")
        diaInicial_obj = dataInicial_obj.day
        mesAnoDataInicial = dataInicial_obj.strftime("%b. de %Y").lower()
        dataInicial_formatada = dataInicial_obj.strftime("%b. de %Y").lower()

        while True:
            mesAnoAtual = driver.find_element(By.XPATH, f'//*[@id="owl-dt-picker-0"]/div[2]/owl-date-time-calendar/div[1]/div').text.lower()
            if dataInicial_formatada == f'{mesAnoAtual}':
                celulasCalendario = driver.find_elements(By.CLASS_NAME, 'owl-dt-calendar-cell-content')
                celulasFiltradas = remover_celulas_anteriores_a_um(celulasCalendario)
                
                for celula in celulasFiltradas:
                    if celula.text == str(diaInicial_obj):
                        celula.click()
                        break
                break
            else:
                dataAtual_obj = mesAnoAtual
                dataInicial
                if f'{dataInicial_formatada}' < dataAtual_obj:
                    driver.find_element(By.XPATH, '//*[@id="owl-dt-picker-0"]/div[2]/owl-date-time-calendar/div[1]/button[2]').click()
                else:
                    driver.find_element(By.XPATH, '//*[@id="owl-dt-picker-0"]/div[2]/owl-date-time-calendar/div[1]/button[1]').click()

        time.sleep(1)

        mesHoje = datetime.now().strftime("%b. de %Y").lower()
        dataFinal_obj = datetime.strptime(str(dataFinal), "%d/%m/%Y")
        diaFinal_obj = dataFinal_obj.day
        dataFinal_formatada = dataFinal_obj.strftime("%b. de %Y").lower()

        while True:
            mesAnoAtual = driver.find_element(By.XPATH, f'//*[@id="owl-dt-picker-0"]/div[2]/owl-date-time-calendar/div[1]/div').text.lower()
            if dataFinal_formatada == f'{mesAnoAtual}':
                calendarioPosSelecao = driver.find_elements(By.CLASS_NAME, 'owl-dt-calendar-cell-content')
                calendarioFiltrado = remover_celulas_anteriores_a_um(calendarioPosSelecao)
                
                for diasPosSelecao in calendarioFiltrado:
                    if diasPosSelecao.text == str(diaFinal_obj):
                        diasPosSelecao.click()
                        break
                break
            else:
                dataAtual_obj = mesAnoAtual
                if f'{dataFinal_formatada}' > dataAtual_obj:
                    driver.find_element(By.XPATH, '//*[@id="owl-dt-picker-0"]/div[2]/owl-date-time-calendar/div[1]/button[2]').click()
                else:
                    driver.find_element(By.XPATH, '//*[@id="owl-dt-picker-0"]/div[2]/owl-date-time-calendar/div[1]/button[1]').click()
            
        
        
        
        
        botaoFiltrar = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="history_filter"]/div[2]/form/div/div[11]/cla-button-text/button')))
        botaoFiltrar.click()
        pbText.value = "Realizando download da planilha"
        pb.value = 0.5
        page.update()
        planilhaDownload = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="export_history"]')))
        planilhaDownload.click()
        botaoPlanilha = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mobile"]/main/app-consumption-history/app-consumption-history-card-title/section/div[2]/ul/li[3]')))
        botaoPlanilha.click()
        time.sleep(5)
        
        def abrir_ultima_planilha_downloads():
            # Caminho para a pasta Downloads
            pasta_downloads = os.path.expanduser("~/Downloads")
            
            # Procurar arquivos CSV
            arquivos_planilhas = glob.glob(os.path.join(pasta_downloads, "*.csv"))
            
            if not arquivos_planilhas:
                return None, None
            
            # Encontrar o arquivo mais recente
            ultima_planilha = max(arquivos_planilhas, key=os.path.getctime)
            
            # Ler a planilha usando pandas
            try:
                df = pd.read_csv(ultima_planilha, delimiter=";")
                
                # Verificar se a coluna 12 (base 0 → índice 11) existe no DataFrame
                if df.shape[1] > 12:
                    # Remover linhas onde a coluna 12 contém "ARLA *"
                    df = df[~df.iloc[:, 12].astype(str).str.contains(r"ARLA\s*\d*", flags=re.IGNORECASE, na=False)]
                
                return df, ultima_planilha
            except Exception as e:
                return None, None

        def criar_ou_atualizar_planilha(df):
            try:
                colunas_desejadas_indices = [2, 1, 4, 7]
                df_selecionado = df.iloc[:, colunas_desejadas_indices].copy()
                df_selecionado.columns = ["PLACA", "DATA", "KM", "LITRAGEM"]
                df_selecionado["PLACA"] = df_selecionado["PLACA"].str.replace("-", "", regex=False)
                df_selecionado.insert(2, "POSTO", "BOHRER")
                df_selecionado.sort_values(by=["PLACA", "DATA"], inplace=True)

                # Converter para numérico, garantindo que valores inválidos sejam convertidos corretamente
                df_selecionado["KM"] = pd.to_numeric(df_selecionado["KM"], errors="coerce")
                df_selecionado["LITRAGEM"] = pd.to_numeric(df_selecionado["LITRAGEM"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
                
                # Agrupar por PLACA, DATA e KM somando os valores de LITRAGEM
                df_selecionado = df_selecionado.groupby(["PLACA", "DATA", "KM", "POSTO"], as_index=False).sum()
                
                placas_unicas = df_selecionado["PLACA"].unique()
                
                area_de_trabalho = os.path.expanduser("~/Desktop")
                caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")
                modo_escrita = "a" if os.path.exists(caminho_arquivo) else "w"

                writer = pd.ExcelWriter(caminho_arquivo, engine="openpyxl", mode=modo_escrita, if_sheet_exists="overlay")

                for placa in placas_unicas:
                    df_placa = df_selecionado[df_selecionado["PLACA"] == placa].copy()
                    df_placa = df_placa.sort_values(by="DATA")

                    if os.path.exists(caminho_arquivo):
                        try:
                            existing_sheets = pd.ExcelFile(caminho_arquivo).sheet_names
                            if placa in existing_sheets:
                                existing_df = pd.read_excel(caminho_arquivo, sheet_name=str(placa))
                                df_placa = pd.concat([existing_df, df_placa], ignore_index=True).drop_duplicates()
                        except Exception as e:
                            print(f"Erro ao ler a planilha existente para a placa {placa}: {e}")

                    df_placa["MEDIA"] = ""
                    df_placa.to_excel(writer, sheet_name=str(placa), index=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets[str(placa)]
                    
                    for row in range(2, len(df_placa) + 2):
                        formula = f'=IF(A{row}<>",",(D{row}-D{row-1})/E{row},"")'
                        worksheet[f"F{row}"] = formula
                
                writer.close()
            except Exception as e:
                print(f"Erro ao criar ou atualizar a planilha: {e}")



        # Delay para garantir que o download da planilha foi concluído
        time.sleep(3)

        # Executar as funções
        df, ultima_planilha = abrir_ultima_planilha_downloads()

        if df is not None:
            pb.value = 0.8
            pbText.value = "Criando/atualizando planilha"
            page.update()
            criar_ou_atualizar_planilha(df)
            pb.value = 1
            pbText.value = "Planilha criada/atualizada com sucesso"
            page.update()
            time.sleep(1)
            pb.value = 0
            pbText.value = "Iniciando processo"
            pbText.visible = False
            pb.visible = False
            page.update()
            
        time.sleep(1)
        driver.quit()
        
    def tradicao(dataInicial, dataFinal):
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.get("https://cliente.argosistemas.com/#/login")
        login = 'manutencao.rede@gmail.com'
        senha = '1945LUCIANO'
        driver.maximize_window()
        wait = WebDriverWait(driver, 500)
    
    def buffon(dataInicial, dataFinal):
        pass
    
    def sander(dataInicial, dataFinal):
        pass
    
    def buscarInformacoes(e):
        if not dataInicial.value or not dataFinal.value or not postos.value:
            snackbar = ft.SnackBar(ft.Text("Por favor, preencha todos os dados"))
            page.overlay.append(snackbar)
            snackbar.open = True
            page.update()
            return
        if postos.value == "S I M":
            sim(dataInicial.value, dataFinal.value)
        elif postos.value == "BOHRER":
            bohrer(dataInicial.value, dataFinal.value)
            pass
        elif postos.value == "TRADIÇÃO":
            #tradicao()
            pass
        elif postos.value == "BUFFON":
            #buffon(dataInicial, dataFinal)
            pass
        elif postos.value == "SANDER":
            #sander(dataInicial, dataFinal)
            pass
        else:
            pass
    def abrirPlanilha(e):
        area_de_trabalho = os.path.expanduser("~/Desktop")
        caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")
        os.startfile(caminho_arquivo)
    botaoAbrirPlanilha = ft.ElevatedButton("Abrir planilha", on_click=abrirPlanilha)
    botaoIniciarProcesso = ft.ElevatedButton(
        "Iniciar processo",
        on_click=buscarInformacoes)
    
    def atualizar_valor_dataInicial(e):
        global datepicker_aberto
        data_obj = datetime.strptime(str(e.control.value), "%Y-%m-%d %H:%M:%S")
        data_formatada = data_obj.strftime("%d/%m/%Y")
        dataInicial.value = data_formatada
        page.close(date_pickerInicial)
        botaoIniciarProcesso.focus()
        datepicker_aberto = False 
        page.update()

    def atualizar_valor_dataFinal(e):
        global datepicker_aberto
        data_obj = datetime.strptime(str(e.control.value), "%Y-%m-%d %H:%M:%S")
        data_formatada = data_obj.strftime("%d/%m/%Y")
        dataFinal.value = data_formatada
        page.close(date_pickerFinal)
        botaoIniciarProcesso.focus()
        datepicker_aberto = False 
        page.update()

    def abrir_date_picker_inicial(e):
        global datepicker_aberto
        if not datepicker_aberto:  # Abrir apenas se ainda não estiver aberto
            datepicker_aberto = True
            page.open(date_pickerInicial)
            
    def abrir_date_picker_final(e):
        global datepicker_aberto
        if not datepicker_aberto:  # Abrir apenas se ainda não estiver aberto
            datepicker_aberto = True
            page.open(date_pickerFinal)
        
    date_pickerInicial = ft.DatePicker(
        on_change=atualizar_valor_dataInicial,
        on_dismiss=lambda _: atualizar_valor_dataInicial,
        date_picker_entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY
    )
    date_pickerFinal = ft.DatePicker(
        on_change=atualizar_valor_dataFinal,
        on_dismiss=lambda _: atualizar_valor_dataInicial,
        date_picker_entry_mode=ft.DatePickerEntryMode.CALENDAR_ONLY
    )
    
    # Adicionar o DatePicker à página
    page.overlay.append(date_pickerInicial)
    page.overlay.append(date_pickerFinal)
    
    
    # Campo de texto com on_focus para abrir o DatePicker
    dataInicial = ft.TextField(
        label="Data Inicial",
        on_focus=abrir_date_picker_inicial
    )
    dataFinal = ft.TextField(
        label="Data final",
        on_focus=abrir_date_picker_final
    )
    datas = ft.Row([dataInicial,dataFinal], ft.MainAxisAlignment.CENTER)
        
    def paginaRelatorio(e):
        page.clean()
        
        #Checkbox TODOS
        def VerificarVerdadeiro(e):
            if checkboxTodos.value == True:
                checkboxSim.value = True
                checkboxBohrer.value = True
                checkboxTradição.value = True
                checkboxBuffon.value = True 
                checkboxSander.value = True
                checkboxShopping.value = True
                checkboxStaTerezinha.value = True
                checkboxTw.value = True
                checkboxNobre.value = True
                checkboxFormulaFlorestal.value = True
                page.update()
            elif checkboxTodos.value == False:
                checkboxSim.value = False
                checkboxBohrer.value = False
                checkboxTradição.value = False
                checkboxBuffon.value = False
                checkboxSander.value = False
                checkboxShopping.value = False  
                checkboxStaTerezinha.value = False
                checkboxTw.value = False
                checkboxNobre.value = False
                checkboxFormulaFlorestal.value = False
                page.update()
                
           
        checkboxTodos = ft.Checkbox(label="MARCAR TODOS", value=False, on_change=VerificarVerdadeiro, label_style=ft.TextStyle(weight=ft.FontWeight.BOLD))
        

        # Checkboxes dos postos
        checkboxSim = ft.Checkbox(label="SIM", value=False)
        checkboxBohrer = ft.Checkbox(label="BOHRER", value=False)
        checkboxTradição = ft.Checkbox(label="TRADIÇÃO", value=False)
        checkboxBuffon = ft.Checkbox(label="BUFFON", value=False)
        checkboxSander = ft.Checkbox(label="SANDER", value=False)
        checkboxShopping = ft.Checkbox(label="SHOPPING CAR", value=False)
        checkboxStaTerezinha = ft.Checkbox(label="SANTA TEREZINHA", value=False)
        checkboxTw = ft.Checkbox(label="TW", value=False)
        checkboxNobre = ft.Checkbox(label="NOBRE", value=False)
        checkboxFormulaFlorestal = ft.Checkbox(label="FORMULA/FLORESTAL", value=False)

        # Inputs para filtro
        inputPlaca = ft.TextField(label="Placa", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False)
        inputDataInicial = ft.TextField(label="Data Inicial (DD/MM/AAAA)", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False )
        inputDataFinal = ft.TextField(label="Data Final (DD/MM/AAAA)", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False)



        # Layout da interface
        linha0 = ft.Row([checkboxTodos], ft.MainAxisAlignment.CENTER)
        linha1 = ft.Row([checkboxSim, checkboxBohrer, checkboxTradição], ft.MainAxisAlignment.CENTER)
        linha2 = ft.Row([checkboxBuffon, checkboxSander, checkboxShopping], ft.MainAxisAlignment.CENTER)
        linha3 = ft.Row([checkboxStaTerezinha, checkboxTw, checkboxNobre, checkboxFormulaFlorestal], ft.MainAxisAlignment.CENTER)
        linha4 = ft.Row([inputDataInicial, inputDataFinal], ft.MainAxisAlignment.CENTER)

        # Função para extrair os dados filtrados
        def extrairExcel(checkboxNobre,checkboxTw ,checkboxSim, checkboxBohrer, checkboxTradição, checkboxBuffon, checkboxSander, checkboxShopping, checkboxStaTerezinha, inputPlaca, inputDataInicial, inputDataFinal):
            xls = pd.ExcelFile(caminho_arquivo)
            inputPlaca = inputPlaca.strip()

            if inputPlaca and inputPlaca in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=inputPlaca)
            else:
                df = pd.concat(pd.read_excel(xls, sheet_name=None), ignore_index=True)

            postos_selecionados = []
            if checkboxSim: postos_selecionados.append("S I M")
            if checkboxBohrer: postos_selecionados.append("BOHRER")
            if checkboxTradição: postos_selecionados.append("TRADIÇÃO")
            if checkboxBuffon: postos_selecionados.append("BUFFON")
            if checkboxSander: postos_selecionados.append("SANDER")
            if checkboxShopping: postos_selecionados.append("SHOPP")
            if checkboxStaTerezinha: postos_selecionados.append("STA")
            if checkboxTw: postos_selecionados.append("TW")
            if checkboxNobre: postos_selecionados.append("NOBRE")
            if checkboxFormulaFlorestal: postos_selecionados.append("FLORESTAL")

            if postos_selecionados:
                df = df[df["POSTO"].isin(postos_selecionados)]
            
            if inputPlaca and "PLACA" in df.columns:
                df = df[df["PLACA"].astype(str).str.contains(inputPlaca, case=False, na=False)]
            
            if "DATA" in df.columns and inputDataInicial and inputDataFinal:
                df["DATA"] = pd.to_datetime(df["DATA"], errors='coerce')
                data_inicial = pd.to_datetime(inputDataInicial, errors='coerce', dayfirst=True)
                data_final = pd.to_datetime(inputDataFinal, errors='coerce', dayfirst=True)
                df = df[(df["DATA"] >= data_inicial) & (df["DATA"] <= data_final)]
                
            print (postos_selecionados)

            caminho_saida = os.path.join(area_de_trabalho, "Relatorio_Filtrado.xlsx")
            df.to_excel(caminho_saida, index=False)

            page.add(ft.Text(f"Relatório gerado com sucesso: {caminho_saida}"))
            os.startfile(caminho_saida)
            
            page.update()

        botaoExtrairRelatorio = ft.ElevatedButton("Extrair relatório", on_click=lambda _: extrairExcel(
            checkboxNobre.value,checkboxTw.value, checkboxSim.value, checkboxBohrer.value, checkboxTradição.value, checkboxBuffon.value, checkboxSander.value, checkboxShopping.value, checkboxStaTerezinha.value, f'{inputPlaca.value}', inputDataInicial.value, inputDataFinal.value))

        def voltar_pagina_inicial(e):
            page.clean()
            main(page)

        botaoVoltar = ft.ElevatedButton("Voltar", on_click=voltar_pagina_inicial)

        linha5 = ft.Row([botaoExtrairRelatorio, botaoVoltar], ft.MainAxisAlignment.CENTER)

        page.add(linha0,linha1, linha2, linha3, inputPlaca, linha4, linha5)
        page.update()

    botaoRelatorio = ft.ElevatedButton("Relatório", on_click=paginaRelatorio)
    
    def inserirManualmente(e):
        page.clean()  
        def alterarTelaComBaseNoPosto(posto):
            page.clean()
            if posto in ['NOBRE', 'TRADIÇÃO', 'TW', 'SANTA TEREZINHA', 'SANDER', 'FLORESTAL/FORMULA']:
                page.add (postos, campoPlaca, campoData, campoKm, campoLitros, botaoInserir, botaoVoltar) 
            elif posto in ['SHOPPING CAR', 'BOHRER', 'BUFFON', 'S I M']:   
                page.add (postos, campoPlaca, campoData,campoLitros, campoKm, botaoInserir, botaoVoltar) 
                
            page.update() 
            
                  
        listaPostos = ["S I M", "BOHRER", "TRADIÇÃO", "BUFFON", "SANDER", "SHOPPING CAR", "SANTA TEREZINHA", "TW", "NOBRE", "FLORESTAL/FORMULA"]
        
        postos = ft.Dropdown(
            label="Postos",
            options=[ft.dropdown.Option(posto) for posto in listaPostos],
            value=listaPostos[0],
            on_change=lambda e: alterarTelaComBaseNoPosto(postos.value),
        )
        campoPlaca = ft.TextField(label="Placa (Sem simbolos)", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False, max_length=7)
        campoLitros = ft.TextField(label="Litros", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False)
        def format_date(e):
            value = e.control.value
            digits = "".join(filter(str.isdigit, value))  # Filtra apenas números
            
            if len(digits) > 8:  # Limita ao máximo de 8 números
                digits = digits[:8]

            formatted_value = ""
            if len(digits) > 0:
                formatted_value += digits[:2]  # Dia
            if len(digits) > 2:
                formatted_value += "/" + digits[2:4]  # Mês
            if len(digits) > 3:
                formatted_value += "/2025"  # Ano fixo

            e.control.value = formatted_value
            e.control.update()
            
        campoData = ft.TextField(label="Data (DD/MM/AAAA)", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False, on_change=format_date)
        campoKm = ft.TextField(label="Km", capitalization=ft.TextCapitalization.CHARACTERS, multiline=False)
        
        def inserir(posto, campoPlaca, campoLitros, campoData, campoKm):
            # Ajuste do nome do posto
            posto = 'STA' if posto == 'SANTA TEREZINHA' else 'FLORESTAL' if posto == 'FLORESTAL/FORMULA' else posto.upper()
            
            # Definição do caminho do arquivo
            area_de_trabalho = os.path.expanduser("~/Desktop")
            caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")
            
            # Captura dos valores
            placa = campoPlaca.value.upper()
            litros = campoLitros.value
            data = campoData.value
            km = campoKm.value

            # Converte a data para datetime (caso não esteja)
            if isinstance(data, str):
                data = datetime.strptime(data, "%d/%m/%Y")

            # Garante que KM seja um número inteiro
            km = int(km)

            # Abrindo a planilha (se não existir, cria)
            if not os.path.exists(caminho_arquivo):
                wb = Workbook()
                ws = wb.active
                ws.title = placa
                ws.append(["PLACA", "DATA", "POSTO", "KM", "LITRAGEM", "MEDIA"])
            else:
                wb = load_workbook(caminho_arquivo)
                if placa in wb.sheetnames:
                    ws = wb[placa]
                else:
                    ws = wb.create_sheet(title=placa)
                    ws.append(["PLACA", "DATA", "POSTO", "KM", "LITRAGEM", "MEDIA"])

            # Verificar se já existe um registro com o mesmo KM
            encontrado = False
            for row in ws.iter_rows(min_row=2):
                if row[3].value == km:  # Coluna KM (D)
                    row[4].value = float(row[4].value) + float(litros)  # Soma a litragem corretamente
                    encontrado = True
                    break

            # Se não encontrou um registro com o mesmo KM, adiciona uma nova linha
            if not encontrado:
                ws.append([placa, data, posto, km, litros, ""])  # Média será recalculada depois
            
            # Ajustar largura das colunas
            colunas = {'A': 9, 'B': 11, 'C': 11, 'D': 11, 'E': 10, 'F': 12}
            for coluna, largura in colunas.items():
                ws.column_dimensions[coluna].width = largura


            # Organizar a planilha por data e km
            organizarPlanilhaPorData(ws)

            # Salvar a planilha
            wb.save(caminho_arquivo)
            wb.close()

        def organizarPlanilhaPorData(ws):
            # Coleta os dados (excluindo o cabeçalho)
            dados = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    data = row[1]
                    if isinstance(data, str):
                        data = datetime.strptime(data, "%d/%m/%Y")  # Converter string para datetime
                    km = int(row[3])  # Garante que KM seja um número inteiro
                    dados.append((row[0], data, row[2], km, row[4], row[5]))
                except ValueError:
                    continue  # Ignorar linhas com erro de conversão

            # Ordena os dados pela segunda coluna (DATA) e, em seguida, pela quarta coluna (KM)
            dados.sort(key=lambda x: (x[1], x[3]))

            # Reescreve os dados ordenados
            for i, row in enumerate(dados, start=2):
                for j, value in enumerate(row, start=1):
                    ws[f"{get_column_letter(j)}{i}"] = value
                    if j == 2:  # Aplicando formatação na coluna de data
                        ws[f"{get_column_letter(j)}{i}"].number_format = "DD/MM/YYYY"

            # Recalcula a média
            for i in range(3, len(dados) + 2):  # Começa da terceira linha
                ws[f"F{i}"] = f'=IF(A{i}<>"",(D{i}-D{i-1})/E{i},"")'
            
            # Limpar os campos
            campoLitros.value = ""
            campoData.value = ""
            campoKm.value = ""
            page.update()
            
            
            
            
        botaoInserir = ft.ElevatedButton("Inserir", on_click=lambda _: inserir(postos.value, campoPlaca, campoLitros, campoData, campoKm))
        def voltar_pagina_inicial(e):
            page.clean()
            main(page)

        botaoVoltar = ft.ElevatedButton("Voltar", on_click=voltar_pagina_inicial)
        page.add(postos, campoPlaca, campoLitros, campoData, campoKm, botaoInserir, botaoVoltar)
     
    def telaImportarXml(e):
        page.clean()
        pbImp = ft.ProgressBar(width=200)
        pbTextImp = ft.Text("Iniciando importação...")
        ColunaPbImp = ft.Column([pbImp, pbTextImp], alignment=ft.alignment.center, horizontal_alignment=ft.CrossAxisAlignment.CENTER, visible=False)
        

        def processar_arquivos(e):
            global caminhoDaPasta
            if e.path:
                caminhoDaPasta = e.path
                texto_caminho.value = f"Pasta selecionada: {e.path}"
                texto_caminho.update()

        # Criando o FilePicker
        seletorDePasta = ft.FilePicker(on_result=processar_arquivos)
        page.overlay.append(seletorDePasta)  # Adicionando ao overlay da página

        # Botão para abrir o seletor de pasta
        botao_seletor = ft.ElevatedButton("Selecionar Pasta", on_click=lambda _: seletorDePasta.get_directory_path())

        # Texto para exibir o caminho da pasta
        texto_caminho = ft.Text("Nenhuma pasta selecionada", color="blue")

        def read_xml_files_from_folder(folder_path):
            if not folder_path:
                return {}

            xml_data_list = {}
            extracted_data = []
            
            ColunaPbImp.visible = True
            pbImp.value = 0
            page.update()
            
            #POSTOS TESTADOS E OK
            #TW
            #SANTA TEREZINHA
            #FLORESTAL
            #TRADICAO
            #NOBRE
            
            
            pbTextImp.value = "Coletando dados dos XMLs..."
            pbImp.value = 0.1
            page.update()
            
            for filename in os.listdir(folder_path):
                if filename.lower().endswith(".xml"):
                    file_path = os.path.join(folder_path, filename)
                    try:
                        with open(file_path, "r", encoding="utf-8") as file:
                            xml_data = xmltodict.parse(file.read())

                            # Extração das informações
                            inf_nfe = xml_data["nfeProc"]["NFe"]["infNFe"]

                            data = inf_nfe["ide"].get("dhEmi", "N/A")

                            # Converte a data se estiver em formato ISO 8601
                            if data != "N/A":
                                try:
                                    data = datetime.fromisoformat(data).strftime("%d/%m/%Y")
                                except ValueError:
                                    data = "Formato Inválido"

                            posto = inf_nfe["emit"].get("xNome", "N/A")
                            
                            produtos = inf_nfe["det"] if isinstance(inf_nfe["det"], list) else [inf_nfe["det"]]
                    
                            for produto in produtos:
                                nome_produto = produto["prod"].get("xProd", "").upper()
                                if "ARLA" in nome_produto:
                                    continue

                            # Obtendo a litragem
                            litragem = (
                                inf_nfe["det"][0]["prod"].get("qCom", "N/A")
                                if isinstance(inf_nfe["det"], list)
                                else inf_nfe["det"]["prod"].get("qCom", "N/A")
                            )

                            # Inicializando as variáveis
                            placa = km = "N/A"

                            # Verifica se a empresa é "COMERCIO DE COMBUSTIVEIS FLORESTAL LTDA"
                            if posto == "COMERCIO DE COMBUSTIVEIS FLORESTAL LTDA":
                                inf_adic = inf_nfe.get("infAdic", {}).get("infCpl", "")

                                if inf_adic:
                                    # Capturar a placa corretamente
                                    match_placa = re.search(r"PLACA:\s*([A-Z0-9]+)", inf_adic, re.IGNORECASE)
                                    if match_placa:
                                        placa = match_placa.group(1).upper()

                                    # Capturar o KM corretamente (evitando capturar "MÉDIA KM")
                                    match_km = re.search(r"\bKM:\s*([\d]+)", inf_adic, re.IGNORECASE)
                                    if match_km:
                                        km = match_km.group(1)
                            else:
                                inf_adic = inf_nfe.get("infAdic", {}).get("infCpl", "")
                                if inf_adic:
                                    for part in inf_adic.split("|"):
                                        part = part.strip()
                                        if "Placa:" in part:
                                            placa = part.split(":")[1].strip().replace("-", "").upper()
                                        elif "Media KM:" in part:
                                            continue  # Ignora "Media KM:" para evitar erro
                                        elif "KM:" in part:  # Só processa "KM:" se "Media KM:" não tiver sido encontrado
                                            km = part.split(":")[1].strip()

                            extracted_data.append({
                                "Placa": placa,
                                "Data": data,
                                "KM": km,
                                "Litragem": litragem,
                                "Posto": posto
                            })
                                            
                        
                    
                    except Exception as e:
                        print(f"Erro ao processar o arquivo {filename}: {e}")

            def inserirExtractedData(data_list):
                def inserir(posto, campoPlaca, campoLitros, campoData, campoKm):
                    # Ajuste do nome do posto
                    if posto.startswith('POSTO SANTA TEREZINHA'):
                        posto = 'STA'
                    elif posto.startswith('COMERCIO DE COMBUSTIVEIS FLORESTAL'):
                        posto = 'FLORESTAL'
                    elif posto.startswith('ABASTECEDORA DE COMBUSTIVEIS T W'):
                        posto = 'TW'
                    elif posto.startswith('NOBRE ABASTECEDORA DE COMBUSTIVEIS'):
                        posto = 'NOBRE'
                    elif posto.startswith('POSTO SHOPPING CAR COMBUSTIVEIS'):
                        posto = 'SHOPPING CAR'
                    elif posto.startswith('MPS COMERCIO DE COMBUSTIVEI'):
                        posto = 'TRADIÇÃO'
                    else:
                        posto = posto.upper()
                    
                    # Definição do caminho do arquivo
                    area_de_trabalho = os.path.expanduser("~/Desktop")
                    caminho_arquivo = os.path.join(area_de_trabalho, "Informacoes combustiveis.xlsx")
                    
                    # Captura dos valores
                    placa = campoPlaca.upper()
                    litros = campoLitros
                    data = campoData
                    km = campoKm

                    # Converte a data para datetime (caso não esteja)
                    if isinstance(data, str):
                        data = datetime.strptime(data, "%d/%m/%Y")

                    # Garante que KM seja um número inteiro
                    km = int(km)

                    # Abrindo a planilha (se não existir, cria)
                    if not os.path.exists(caminho_arquivo):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = placa
                        ws.append(["PLACA", "DATA", "POSTO", "KM", "LITRAGEM", "MEDIA"])
                    else:
                        wb = load_workbook(caminho_arquivo)
                        if placa in wb.sheetnames:
                            ws = wb[placa]
                        else:
                            ws = wb.create_sheet(title=placa)
                            ws.append(["PLACA", "DATA", "POSTO", "KM", "LITRAGEM", "MEDIA"])

                    # Verificar se já existe um registro com o mesmo KM
                    encontrado = False
                    for row in ws.iter_rows(min_row=2):
                        if row[3].value == km:  # Coluna KM (D)
                            row[4].value = float(row[4].value) + float(litros)  # Soma a litragem corretamente
                            encontrado = True
                            break

                    # Se não encontrou um registro com o mesmo KM, adiciona uma nova linha
                    if not encontrado:
                        ws.append([placa, data, posto, km, litros, ""])  # Média será recalculada depois
                    
                    # Ajustar largura das colunas
                    colunas = {'A': 9, 'B': 11, 'C': 11, 'D': 11, 'E': 10, 'F': 12}
                    for coluna, largura in colunas.items():
                        ws.column_dimensions[coluna].width = largura

                    pbTextImp.value = "Organizando planilhas..."
                    pbImp.value = 0.7
                    page.update()
                    # Organizar a planilha por data e km
                    organizarPlanilhaPorData(ws)
                    pbTextImp.value = "Salvando planilha"
                    pbImp.value = 0.9
                    page.update()
                    # Salvar a planilha
                    wb.save(caminho_arquivo)
                    wb.close()
                    pbTextImp.value = "Importação concluida!"
                    pbImp.value = 1
                    ColunaPbImp.visible = False
                    page.update()

                def organizarPlanilhaPorData(ws):
                    # Coleta os dados (excluindo o cabeçalho)
                    dados = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        try:
                            data = row[1]
                            if isinstance(data, str):
                                data = datetime.strptime(data, "%d/%m/%Y")  # Converter string para datetime
                            km = int(row[3])  # Garante que KM seja um número inteiro
                            dados.append((row[0], data, row[2], km, row[4], row[5]))
                        except ValueError:
                            continue  # Ignorar linhas com erro de conversão

                    # Ordena os dados pela segunda coluna (DATA) e, em seguida, pela quarta coluna (KM)
                    dados.sort(key=lambda x: (x[1], x[3]))

                    # Reescreve os dados ordenados
                    for i, row in enumerate(dados, start=2):
                        for j, value in enumerate(row, start=1):
                            ws[f"{get_column_letter(j)}{i}"] = value
                            if j == 2:  # Aplicando formatação na coluna de data
                                ws[f"{get_column_letter(j)}{i}"].number_format = "DD/MM/YYYY"

                    # Recalcula a média
                    for i in range(3, len(dados) + 2):  # Começa da terceira linha
                        ws[f"F{i}"] = f'=IF(A{i}<>"",(D{i}-D{i-1})/E{i},"")'
                        
                pbTextImp.value = "Inserindo dados na planilha"
                pbImp.value = 0.5
                for data in data_list:
                    inserir(data["Posto"], data["Placa"], data["Litragem"], data["Data"], data["KM"])
            
            inserirExtractedData(extracted_data)    
        # Botão para iniciar a importação
        botaoIniciarImportacao = ft.ElevatedButton("Iniciar Importação", on_click=lambda _: read_xml_files_from_folder(caminhoDaPasta))
        def voltar_pagina_inicial(e):
            page.clean()
            main(page)

        botaoVoltar = ft.ElevatedButton("Voltar", on_click=voltar_pagina_inicial)

        # Adicionando os elementos à página
        page.add(botao_seletor, texto_caminho, botaoIniciarImportacao,ColunaPbImp, botaoVoltar)
        
        
        
        
        
        
        
    botaoMenuInserir = ft.ElevatedButton("Inserir manualmente", on_click=inserirManualmente,)
    textoRelatorio = ft.Text("Manusear planilha", weight=ft.FontWeight.BOLD,size=20)
    textoInsercao = ft.Text("Inserção De Dados", weight=ft.FontWeight.BOLD,size=20)
    botaoImportarXml = ft.ElevatedButton("Importar XML", on_click=telaImportarXml,)
    
    page.add(postos,datas,
             
            ft.Container(
            content=botaoIniciarProcesso,
            alignment=ft.alignment.center,
            expand=False,
            height=50,
        ),ColunaPb,
            ft.Row(
                controls=[
            ft.Container(content=ft.Column(controls=[textoInsercao,botaoMenuInserir, botaoImportarXml],horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.START, expand=False),border=ft.border.all(1, "black"), alignment=ft.alignment.top_center, expand=False, height=130, width=200),
            
            
            ft.Container(ft.Container(content=ft.Column(controls=[textoRelatorio, botaoAbrirPlanilha, botaoRelatorio], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.alignment.bottom_center),
                                      alignment=ft.alignment.bottom_center, expand=False, bgcolor="light grey",height=200,width=200,border=ft.border.all(1, "black")), 
                         alignment=ft.alignment.bottom_right,height=130, expand=False, margin=ft.margin.only(top=50)),
            
            ],vertical_alignment=ft.CrossAxisAlignment.END, expand=True)
            
            
            )
    
    
    page.update()
    
if __name__ == "__main__":
    ft.app(target=main)
